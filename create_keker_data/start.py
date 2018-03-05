from core import *
from openpyxl import load_workbook
import os
import json
import datetime

class Main_core(object):
    keke_url = 'http://test.api.keker.xincheng.tv'
    workbook = load_workbook(filename="test_data.xlsx")

    def __init__(self, conn = None):
        self.conn = conn

    @classmethod
    def excel_sheet_user(cls, sheet):
        user_list = []
        for i in range(2, sheet.max_row + 1):
            user={}
            user['phone'] = str(sheet.cell(row=i, column=1).value)
            user['password'] = str(sheet.cell(row=i, column=2).value)
            user['success'] = 0

            user_list.append(user)
        return user_list

    @classmethod
    def excel_sheet_card(cls, sheet):
        card_list = []
        for i in range(2, sheet.max_row + 1):
            card={}
            card['phone'] = str(sheet.cell(row=i, column=1).value)
            card['name'] = str(sheet.cell(row=i, column=2).value)
            card['alias'] = str(sheet.cell(row=i, column=3).value)
            card['email'] = str(sheet.cell(row=i, column=4).value)
            card['country'] = str(sheet.cell(row=i, column=5).value)
            card['province'] = str(sheet.cell(row=i, column=6).value)
            card['city'] = str(sheet.cell(row=i, column=7).value)
            card['reg_type'] = str(sheet.cell(row=i, column=8).value)
            card['card_type'] = str(sheet.cell(row=i, column=9).value)
            card['company'] = str(sheet.cell(row=i, column=10).value)
            card['com_fullname'] = str(sheet.cell(row=i, column=11).value)
            card['com_address'] = str(sheet.cell(row=i, column=12).value)
            card['com_country'] = str(sheet.cell(row=i, column=13).value)
            card['com_province'] = str(sheet.cell(row=i, column=14).value)
            card['com_city'] = str(sheet.cell(row=i, column=15).value)
            card['com_town'] = str(sheet.cell(row=i, column=16).value)
            card['com_website'] = str(sheet.cell(row=i, column=17).value)
            card['com_department'] = str(sheet.cell(row=i, column=18).value)
            card['com_position'] = str(sheet.cell(row=i, column=19).value)
            card['com_entry'] = str(sheet.cell(row=i, column=20).value)
            card['com_leave'] = str(sheet.cell(row=i, column=21).value)
            card['school'] = str(sheet.cell(row=i, column=22).value)
            card['sch_country'] = str(sheet.cell(row=i, column=23).value)
            card['sch_province'] = str(sheet.cell(row=i, column=24).value)
            card['sch_city'] = str(sheet.cell(row=i, column=25).value)
            card['sch_town'] = str(sheet.cell(row=i, column=26).value)
            card['sch_department'] = str(sheet.cell(row=i, column=27).value)
            card['sch_major'] = str(sheet.cell(row=i, column=28).value)
            card['sch_class'] = str(sheet.cell(row=i, column=29).value)
            card['sch_drom'] = str(sheet.cell(row=i, column=30).value)
            card['sch_graduation'] = str(sheet.cell(row=i, column=31).value)
            card['agent_company'] = str(sheet.cell(row=i, column=32).value)
            card['fields'] = str(sheet.cell(row=i, column=33).value)
            card['agent_phone'] = str(sheet.cell(row=i, column=34).value)
            card['agent_name'] = str(sheet.cell(row=i, column=35).value)
            card['avatar'] = str(sheet.cell(row=i, column=36).value)
            card['success'] = 0
            card_list.append(card)
        return card_list

    @classmethod
    def excel_sheet_flow(cls, sheet):
        flow_list = []
        for i in range(2, sheet.max_row + 1):
            flow={}
            flow['phone'] = str(sheet.cell(row=i, column=1).value)
            friends = str(sheet.cell(row=i, column=2).value)
            friends = friends.split('|')
            flow['friends'] = friends
            flow['success'] = 0
            flow_list.append(flow)
        return flow_list

    @classmethod
    def excel_sheet_schedule_p(cls, sheet):
        schedule_p_list = []
        for i in range(2, sheet.max_row + 1):
            schedule_p={}
            schedule_p['phone'] = str(sheet.cell(row=i, column=1).value)
            schedule_p['proxy'] = str(sheet.cell(row=i, column=2).value)
            schedule_p['hour'] = sheet.cell(row=i, column=3).value
            schedule_p['country'] = str(sheet.cell(row=i, column=4).value)
            schedule_p['province'] = str(sheet.cell(row=i, column=5).value)
            schedule_p['city'] = str(sheet.cell(row=i, column=6).value)
            schedule_p['town'] = str(sheet.cell(row=i, column=7).value)
            schedule_p['location'] = str(sheet.cell(row=i, column=8).value)
            schedule_p['theme'] = str(sheet.cell(row=i, column=9).value)
            schedule_p['public'] = str(sheet.cell(row=i, column=10).value)
            schedule_p['success'] = 0
            schedule_p_list.append(schedule_p)
        return schedule_p_list

    @classmethod
    def excel_sheet_schedule_t(cls, sheet):
        schedule_t_list = []
        for i in range(2, sheet.max_row + 1):
            schedule_t = {}
            schedule_t['phone'] = str(sheet.cell(row=i, column=1).value)
            schedule_t['type'] = sheet.cell(row=i, column=2).value
            schedule_t['day'] = sheet.cell(row=i, column=3).value
            schedule_t['country'] = sheet.cell(row=i, column=4).value
            schedule_t['province'] = sheet.cell(row=i, column=5).value
            schedule_t['city'] = sheet.cell(row=i, column=6).value
            schedule_t['location'] = str(sheet.cell(row=i, column=7).value)
            schedule_t['theme'] = str(sheet.cell(row=i, column=8).value)
            schedule_t['desc'] = str(sheet.cell(row=i, column=9).value)
            schedule_t['public'] = str(sheet.cell(row=i, column=10).value)
            schedule_t['success'] = 0
            schedule_t_list.append(schedule_t)
        return schedule_t_list

    @classmethod
    def excel_sheet_schedule_k(cls, sheet):
        schedule_k_list = []
        for i in range(2, sheet.max_row + 1):
            schedule_k={}
            schedule_k['phone'] = str(sheet.cell(row=i, column=1).value)
            schedule_k['proxy'] = sheet.cell(row=i, column=2).value
            schedule_k['hour'] = sheet.cell(row=i, column=3).value
            schedule_k['country'] = str(sheet.cell(row=i, column=4).value)
            schedule_k['province'] = str(sheet.cell(row=i, column=5).value)
            schedule_k['city'] = str(sheet.cell(row=i, column=6).value)
            schedule_k['town'] = str(sheet.cell(row=i, column=7).value)
            schedule_k['location'] = str(sheet.cell(row=i, column=8).value)
            schedule_k['theme'] = str(sheet.cell(row=i, column=9).value)
            schedule_k['desc'] = str(sheet.cell(row=i, column=10).value)
            schedule_k['public'] = str(sheet.cell(row=i, column=11).value)
            schedule_k['allow'] = str(sheet.cell(row=i, column=12).value)
            schedule_k['max_num'] = sheet.cell(row=i, column=13).value
            schedule_k['fee'] = str(sheet.cell(row=i, column=14).value)
            schedule_k['success'] = 0
            schedule_k_list.append(schedule_k)
        return schedule_k_list



    def keker_register(self, phone, password):
        password = hashlib.md5(password.encode(encoding='utf-8'))
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        keke_url = Main_core.keke_url
        url = keke_url + '/v1/portal/register'
        sms_url = keke_url + '/v1/sms'
        login_url = keke_url + '/v1/portal/login'
        headers = ''
        cookies = ''

        user_keke = sql_core.get_user_by_phone(phone)
        if user_keke is None:
            data = {}
            data['phone'] = phone
            data['passwd'] = password.hexdigest()
            data['client'] = 'android'
            data['device'] = password.hexdigest()

            data_sms = {}
            data_sms['phone'] = data['phone']
            data_sms['behavior'] = 'register'
            payload_sms = data_sms

            #发短信
            sms_data = sql_core.post_request(sms_url, payload_sms, headers, cookies)

            if sms_data.status_code == 200:
                sms_data_josn = sms_data.json()
                if sms_data_josn['status'] == '200':
                    data['smskey'] = sms_data_josn['message']['smskey']
                    num = sql_core.get_sms_number_by_phone(data['phone'], data_sms['behavior'])
                    if num is None:
                        return None
                    data['smscode'] = num

            if data['smscode'] == '' or data['smskey'] == '':
                print('短信验证失败'+ phone)

            payload = data
            #执行注册
            r = sql_core.post_request(url, payload, headers, cookies)
            if r.status_code == 200:
                if r.json()['status'] != '200':
                    print("注册失败: "+ r.json() + phone)

            user_keke = sql_core.get_user_by_phone(phone)

        request_data = {}
        request_data['id'] = user_keke[0]
        request_data['token'] = ''

        print("执行登录: " + phone)
        data_login = {}
        data_login['phone'] = phone
        data_login['passwd'] = password.hexdigest()
        data_login['device'] = password.hexdigest()
        data_login['client'] = 'android'
        payload_login = data_login

        login_data = sql_core.post_request(login_url, payload_login, headers, cookies)

        if login_data.status_code == 200:
            login_json = login_data.json()
            print(login_json)
            if login_json['status'] == '200':
                json_data = {}
                json_data['auth_session'] = login_json['message']['code']
                json_data['sid'] = login_json['message']['sid']
                json_data['token'] = login_json['message']['token']
                authorization = sql_core.create_authorization(json_data['auth_session'], 'android_id', json_data['sid'],json_data['token'], 'KEKER_ANDROID')
                request_data['token'] = authorization
            else:
                print(login_json)
        conn.close()
        return request_data

    def keke_update_userinfo(self, user, token):
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        keke_url = Main_core.keke_url
        city_url = keke_url + '/v1/user/profile/city'
        add_card_url = keke_url + '/v1/user/card/add'
        fieldlist_url = keke_url + '/v1/common/act/fieldlist'
        profile_type_url = keke_url + '/v1/user/profile/type'

        buddy_location = keke_url + '/v1/buddy/set/location'


        headers = {'Authorization':token}
        cookies = ''

        city_code_data = {}
        city_code_data['city_code'] = user['city']
        payload_city_code= city_code_data
        r = sql_core.post_request(buddy_location, payload_city_code, headers, cookies)
        if r.status_code == 200:
            if r.json()['status'] != '200':
                print("更新当前地理位置(好友失败): "+user['phone'])
                print(r.json())
        city = {}
        city['country'] = user['country']
        city['province'] = user['province']
        city['city'] = user['city']
        payload_city = city

        r = sql_core.post_request(city_url, payload_city, headers, cookies)

        if r.status_code == 200:
            if r.json()['status'] != '200':
                print("设置城市失败: "+user['phone'])
                print(r.json())

        uploads_url = keke_url + '/v1/common/uploads'
        img_path = os.getcwd() +'/img/'+str(user['avatar']) +'.png'


        data = {}
        data['avatar'] = ''
        files = {
            "file[]": ('temp.jpeg',open(img_path, "rb"),'image/jpeg')
        }
        imgs = {}
        img_request = sql_core.post_request_file(uploads_url, imgs, files, headers, cookies)
        if img_request.status_code == 200:
            if img_request.json()['status'] == '200':
                data['avatar'] = img_request.json()['message'][0]

        if data['avatar'] == '':
            print("头像上传你失败: "+user['phone'])
            print(img_request.text)

        edit_userinfo_url = keke_url + '/v1/user/profile/edit'
        data['name'] = user['name']
        data['alias'] = user['alias']
        data['email'] = user['email']
        data['country'] = user['country']
        data['province'] = user['province']
        data['city'] = user['city']
        payload = data
        r = sql_core.post_request(edit_userinfo_url, payload, headers, cookies)
        if r.status_code == 200:
            if r.json()['status'] != '200':
                print("设置用户信息失败: "+user['phone'])
                print(r.json())

        profile_type_data = {}
        card_data = {}
        card_data['card_type'] = user['card_type']
        card_data['reg_type'] = user['reg_type']
        card_data['if_card'] = '0'

        if user['card_type'] == 'worker':
            profile_type_data['idtype'] = 'w'
            card_data['company'] = user['company']
            likelist_url = keke_url + '/v1/common/company/likelist?company_name='+ user['company']
            r = sql_core.get_request(likelist_url, '', headers, cookies)
            if r.status_code == 200:
                print(r.json())
                if r.json()['status'] == '200':
                    card_data['company'] = r.json()['company_list'][0]['id']

            card_data['com_fullname'] = user['com_fullname']
            card_data['com_address'] = user['com_address']
            card_data['com_country'] = user['com_country']
            card_data['com_province'] = user['com_province']
            card_data['com_city'] = user['com_city']
            card_data['com_website'] = user['com_website']
            card_data['com_department'] = user['com_department']
            card_data['com_position'] = user['com_position']
            card_data['com_entry'] = user['com_entry']
            card_data['com_leave'] = user['com_leave']

        if user['card_type'] == 'student':
            profile_type_data['idtype'] = 's'

            school_url = keke_url + '/v1/common/school/likelist?school_name='+user['school']
            r = sql_core.get_request(school_url, '', headers, cookies)

            if r.status_code == 200:
                print(r.json())
                if r.json()['status'] == '200':
                    card_data['school'] = r.json()['school_list'][0]['id']
                else:
                    card_data['school'] = user['school']

            card_data['sch_country'] = user['sch_country']
            card_data['sch_province'] = user['sch_province']
            card_data['sch_city'] = user['sch_city']
            card_data['sch_department'] = user['sch_department']
            card_data['sch_major'] = user['sch_major']
            card_data['sch_class'] = user['sch_class']
            card_data['sch_graduation'] = user['sch_graduation']

        if user['card_type'] == 'artist':
            profile_type_data['idtype'] = 'e'
            card_data['agent_company'] = user['agent_company']

            likelist_url = keke_url + '/v1/common/company/likelist?company_name='+ user['agent_company']
            r = sql_core.get_request(likelist_url, '', headers, cookies)
            if r.status_code == 200:
                print(r.json())
                if r.json()['status'] == '200':
                    card_data['agent_company'] = r.json()['company_list'][0]['id']

            fieldlist_data = sql_core.get_request(fieldlist_url,'',headers,cookies)
            if fieldlist_data.status_code == 200:
                fieldlist_data_dict = fieldlist_data.json()
                fields = user['fields'].split('|')
                fields_list = []
                fieldlist_data_dict = fieldlist_data_dict['data']
                for i in range(len(fields)):
                    for j in range(len(fieldlist_data_dict)):
                        if fieldlist_data_dict[j]['name'] == fields[i]:
                            fields_list.append(fieldlist_data_dict[j]['id'])
            field_data = json.dumps(fields_list, ensure_ascii=False)
            card_data['fields'] = field_data
            card_data['agent_phone'] = user['agent_phone']
            card_data['agent_name'] = user['agent_name']

        payload = profile_type_data
        r = sql_core.post_request(profile_type_url, payload, headers, cookies)
        if r.status_code ==200:
            if r.json()['status'] != '200':
                print("设置用户职业失败: "+user['phone'])
                print(r.json())

        payload = card_data
        print(payload)
        r = sql_core.post_request(add_card_url, payload, headers, cookies)
        if r.status_code ==200:
            if r.json()['status'] != '200':
                print("增加用户卡片失败: "+user['phone'])
                print(r.json())


    def keke_add_buddy(self,token,uids):
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        keke_url = Main_core.keke_url
        news_buddy_url = keke_url+'/v1/buddy/new'
        operation_url = keke_url+'/v1/buddy/new_buddy'

        add_buddy_url = keke_url+'/v1/buddy/apply/add'
        headers = {'Authorization':token}
        cookies = ''


        #查询是否有新好友
        r = sql_core.get_request(news_buddy_url,'',headers,cookies)
        if r.status_code==200:
            new_buddy_data = r.json()
            if new_buddy_data['status'] == '200':
                #解析出uid
                buddy_datas = new_buddy_data['message']
                for buddy_data in buddy_datas:
                    data = {}
                    data['status'] = 1
                    data['user_id'] = buddy_data['user_id']
                    payload = data
                    r = sql_core.post_request(operation_url, payload, headers, cookies)
                    if r.status_code == 200:
                        r_status = r.json()
                        if r_status['status'] != '200':
                            print('操作好友失败'+r.text)
        else:
            print('好友列表获取失败')
        #添加好友
        for uid in uids:
            data = {}
            data['user_id'] = uid
            payload = data
            r = sql_core.post_request(add_buddy_url, payload, headers, cookies)
            if r.status_code != 200:
                print('添加好友失败')

    #个人
    def add_schedule_p(self, token, schedule_p):
        keke_url = mcore.keke_url
        url = keke_url + '/v1/private/add'
        headers = {'Authorization':token}
        cookies = ''
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        start_time = int(time.time())
        time_day = str(time.strftime('%Y-%m-%d')) + ' 23:59:59'
        timeArray = time.strptime(time_day, "%Y-%m-%d %H:%M:%S")
        end_time = int(time.mktime(timeArray))
        z_end_time = start_time + (3600 * schedule_p['hour'])

        if z_end_time >= end_time:
            end_time = end_time

        data = {}
        data['begtime'] = start_time
        data['endtime'] = end_time
        data['country'] = schedule_p['country']
        data['province'] = schedule_p['province']
        data['city'] = schedule_p['city']
        data['town'] = schedule_p['town']
        data['location'] = schedule_p['location']
        data['theme'] = schedule_p['theme']
        data['public'] = schedule_p['public']
        data['proxy'] = schedule_p['proxy']
        payload = data
        print(payload)
        r = sql_core.post_request(url, payload, headers, cookies)
        if r.status_code == 200:
            r_data = r.json()
            if r_data['status'] != '200':
                print('创建个人档期失败')
                print(r.json())
        else:
            print('创建个人档期失败')
            print(r.text)

    #差旅
    def add_schedule_t(self, token,schedule_t):
        keke_url = mcore.keke_url
        url = keke_url + '/v1/schedule/travel/add'
        headers = {'Authorization':token}
        cookies = ''
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        today = datetime.date.today()
        list_date = []
        for i in range(0, schedule_t['day']):
            tomorrow = today + datetime.timedelta(days=i)
            list_date.append(str(tomorrow))

        list_date_json = json.dumps(list_date, ensure_ascii=False).encode('utf-8')

        data = {}
        data['type'] = schedule_t['type']
        data['date'] = list_date_json
        data['country'] = schedule_t['country']
        data['province'] = schedule_t['province']
        data['public'] = schedule_t['public']
        data['city'] = schedule_t['city']
        data['location'] = schedule_t['location']
        data['theme'] = schedule_t['theme']
        data['desc'] = schedule_t['desc']
        payload = data
        print(payload)
        r = sql_core.post_request(url, payload, headers, cookies)
        if r.status_code == 200:
            print(r.json())
            if r.json()['status'] != '200':
                print("添加差旅档期失败: " + r.text)

    #科科
    def add_schedule_k(self, token,schedule_k):
        keke_url = mcore.keke_url
        url = keke_url + '/v1/keke/add'
        headers = {'Authorization': token}
        cookies = ''
        conn = get_conn_keke()
        sql_core = Sql_Core(conn)

        start_time = int(time.time())
        time_day = str(time.strftime('%Y-%m-%d')) + ' 23:59:00'
        timeArray = time.strptime(time_day, "%Y-%m-%d %H:%M:%S")
        end_time = int(time.mktime(timeArray))
        z_end_time = start_time + (3600 * schedule_k['hour'])

        if z_end_time >= end_time:
            end_time = end_time

        data = {}
        data['begtime'] = start_time
        data['endtime'] = end_time
        data['country'] = schedule_k['country']
        data['province'] = schedule_k['province']
        data['city'] = schedule_k['city']
        data['town'] = schedule_k['town']

        data['location'] = schedule_k['location']
        data['theme'] = schedule_k['theme']
        data['public'] = schedule_k['public']
        data['proxy'] = schedule_k['proxy']
        data['desc'] = schedule_k['desc']
        data['allow'] = schedule_k['allow']
        data['max_num'] = schedule_k['max_num']
        data['fee'] = schedule_k['fee']

        #来源文档没有该参数 需要增加
        data['ref_type'] = 's' #schedule_k['ref_type']
        data['ref_id'] = 0



        payload = data
        print(payload)
        r = sql_core.post_request(url, payload, headers, cookies)
        if r.status_code == 200:
            print(r.json())
            r_data = r.json()
            if r_data['status'] != '200':
                print('创建科科档期失败')
                print(r.json())
        else:
            print('创建科科档期失败')
            print(r.text)

    def get_excel_by_sheet_name(self,sheet_name, workbook):
        data = ''
        worksheets = workbook.sheetnames
        run = False
        for worksheet in worksheets:
            if worksheet == sheet_name:
                run = True
            if run:
                break
        if run:
            sheet = workbook[sheet_name]

            if sheet_name == 'user':
                data = Main_core.excel_sheet_user(sheet)
                return data
            if sheet_name == 'card':
                data = Main_core.excel_sheet_card(sheet)
                return data
            if sheet_name == 'flow':
                data = Main_core.excel_sheet_flow(sheet)
                return data
            if sheet_name == 'schedule_p':
                data = Main_core.excel_sheet_schedule_p(sheet)
                return data
            if sheet_name == 'schedule_t':
                data = Main_core.excel_sheet_schedule_t(sheet)
                return data
            if sheet_name == 'schedule_k':
                data = Main_core.excel_sheet_schedule_k(sheet)
                return data

if __name__ == "__main__":
    conn = ''
    while(True):
        try:
            conn = get_conn_keke()
        except BaseException as e:
            print('异常错误' + str(e))
        if conn is not None:
            break

    mcore = Main_core(conn)
    workbook = Main_core.workbook

    user_list = mcore.get_excel_by_sheet_name('user', workbook)
    user_tokens = {}


    #执行注册拿到 token
    user_num = len(user_list)
    while(True):
        for i in range(0,len(user_list)):
            try:
                print('d: '+ str(i) +'  z: '+ str(user_num))
                user = user_list[i]

                if user['success'] == 0:
                    register_data = mcore.keker_register(user['phone'], user['password'])
                    tmp_data = {}
                    tmp_data['token'] = register_data['token']
                    tmp_data['id'] = register_data['id']
                    user_tokens[user['phone']] = tmp_data
                    user_num = user_num -1
                    user['success'] = 1
                    user_list[i] = user
            except BaseException as e:
                print('登录异常错误'+ str(e) +'|'+ str(user['phone']))
        if user_num == 0:
            break

       # time.sleep(1)
    print(user_tokens)

    #完善跟人资料 创建电子名片
    card_list = mcore.get_excel_by_sheet_name('card', workbook)
    print(card_list)
    card_success_num = len(card_list)
    while(True):
        for i in range(0,len(card_list)):
            try:
                card = card_list[i]
                if card['success'] == 0:
                    user = user_tokens[card['phone']]
                    mcore.keke_update_userinfo(card, user['token'])
                    card_success_num = card_success_num - 1
                    card['success'] = 1
                    card_list[i] = card
            except BaseException as e:
                print('异常错误_创建电子名片' + str(e) + '|' + str(card['phone']))
        if card_success_num == 0:
            break

    #添加好友关系
    flow_list = mcore.get_excel_by_sheet_name('flow', workbook)
    print(flow_list)
    flow_success_num = len(flow_list)
    while(True):
        for i in range(0,len(flow_list)):
            try:
                flow = flow_list[i]
                if flow['success'] == 0:
                    user = user_tokens[flow['phone']]
                    friends_uid = []
                    friends = flow['friends']
                    for j in range(len(friends)):
                        try:
                            if friends[j] != 'None':
                                uid = user_tokens[friends[j]]['id']
                                friends_uid.append(uid)
                        except BaseException as e:
                            print('获取uid失败' + str(e) + '|' + str(friends[j]))
                    mcore.keke_add_buddy(user['token'], friends_uid)
                    flow_success_num = flow_success_num - 1
                    flow['success'] = 1
                    flow_list[i] = flow
            except BaseException as e:
                print('好友关系异常错误' + str(e) + '|' + str(flow['phone']))
        if flow_success_num == 0:
            break

    #添加个人档期
    schedule_p_list = mcore.get_excel_by_sheet_name('schedule_p', workbook)
    print(schedule_p_list)
    schedule_p_success_num = len(schedule_p_list)
    while(True):
        for i in range(len(schedule_p_list)):
            schedule_p = schedule_p_list[i]
            try:
                if schedule_p['success'] == 0:
                    phone = schedule_p['phone']
                    token = user_tokens[phone]['token']
                    mcore.add_schedule_p(token, schedule_p)
                    schedule_p_success_num = schedule_p_success_num - 1
                    schedule_p['success'] = 1
                    schedule_p_list[i] = schedule_p
            except BaseException as e:
                print('个人档期异常错误' + str(e) + '|' + phone)
        if schedule_p_success_num == 0:
            break


    #添加差旅档期
    schedule_t_list = mcore.get_excel_by_sheet_name('schedule_t', workbook)
    print(schedule_t_list)
    schedule_t_success_num = len(schedule_t_list)
    while(True):
        for i in range(len(schedule_t_list)):
            schedule_t = schedule_t_list[i]
            try:
                if schedule_t['success'] == 0:
                    phone = schedule_t['phone']
                    token = user_tokens[phone]['token']
                    mcore.add_schedule_t(token, schedule_t)
                    schedule_t_success_num = schedule_t_success_num -1
                    schedule_t['success'] = 1
                    schedule_t_list[i] = schedule_t
            except BaseException as e:
                print('差旅异常错误' + str(e) + '|' + phone)
        if schedule_t_success_num == 0:
            break

    #添加科科档期
    schedule_k_list = mcore.get_excel_by_sheet_name('schedule_k', workbook)
    print(schedule_k_list)

    schedule_k_success_num = len(schedule_k_list)

    while(True):
        for i in range(len(schedule_k_list)):
            schedule_k = schedule_k_list[i]
            try:
                if schedule_k['success'] == 0:
                    phone = schedule_k['phone']
                    token = user_tokens[phone]['token']
                    mcore.add_schedule_k(token, schedule_k)
                    schedule_k_success_num = schedule_k_success_num - 1
                    schedule_k['success'] = 1
                    schedule_k_list[i] = schedule_k
            except BaseException as e:
                print('科科异常错误' + str(e) + '|' + phone)

        if schedule_k_success_num == 0:
            break

    # if conn:
    #     conn.close()